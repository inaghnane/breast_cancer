def get_results():
    import random
    import numpy
    import pandas

    random.seed(0)
    numpy.random.seed(0)

    dataframe = pandas.read_csv("../Breast Cancer Wisconsin.csv")
    dataframe = dataframe.iloc[:, 1:]

    #* outliers
    from sklearn.ensemble import IsolationForest
    X = dataframe.iloc[:, 1:] 
    y = dataframe.iloc[:, 0] 
    model = IsolationForest(contamination=0.05)
    outliers = model.fit_predict(X)

    #* standard scaling
    from sklearn.preprocessing import StandardScaler
    standard_scaler = StandardScaler()
    standard_scaler.fit_transform(X=X)

    #* splitting data
    from sklearn.model_selection import train_test_split
    X_train, X_test, y_train, y_test = train_test_split(X, y, train_size=0.85)

    #* LogisticRegression model
    from sklearn.linear_model import LogisticRegression
    best_log_model = LogisticRegression(solver='liblinear', C=10.0, random_state=0, max_iter=100)
    best_log_model.fit(X_train, y_train)

    #* K-NN model
    from sklearn.neighbors import KNeighborsClassifier
    best_knn_model = KNeighborsClassifier(n_neighbors=6)
    best_knn_model.fit(X_train, y_train)

    #* DecisionTreeClassifier model
    from sklearn.tree import DecisionTreeClassifier
    best_tree_model = DecisionTreeClassifier(max_depth=3)
    best_tree_model.fit(X_train, y_train)


    #* 7. Déploiement et Utilisation (optionnel)
    import openpyxl
    import openpyxl.utils
    import csv 

    def french2english(word):
        translated_names = {
            "Rayon moyen"                  : "radius_mean",
            "Texture moyenne"              : "texture_mean",
            "Périmètre moyen"              : "perimeter_mean",
            "Aire moyenne"                 : "area_mean",
            "Lissé moyen"                  : "smoothness_mean",
            "Compacité moyenne"            : "compactness_mean",
            "Concavité moyenne"            : "concavity_mean",
            "Points concaves moyens"       : "concave_points_mean",
            "Symétrie moyenne"             : "symmetry_mean",
            "Dimension fractale moyenne"   : "fractal_dimension_mean",
            "Écart type rayon"             : "radius_se",
            "Écart type texture"           : "texture_se",
            "Écart type périmètre"         : "perimeter_se",
            "Écart type aire"              : "area_se",
            "Écart type lissé"             : "smoothness_se",
            "Écart type compacité"         : "compactness_se",
            "Écart type concavité"         : "concavity_se",
            "Écart type points concaves"   : "concave_points_se",
            "Écart type symétrie"          : "symmetry_se",
            "Écart type dimension fractale": "fractal_dimension_se",
            "Rayon pire"                   : "radius_worst",
            "Texture pire"                 : "texture_worst",
            "Périmètre pire"               : "perimeter_worst",
            "Aire pire"                    : "area_worst",
            "Lissé pire"                   : "smoothness_worst",
            "Compacité pire"               : "compactness_worst",
            "Concavité pire"               : "concavity_worst",
            "Points concaves pires"        : "concave_points_worst",
            "Symétrie pire"                : "symmetry_worst",
            "Dimension fractale pire"      : "fractal_dimension_worst"
        } 
        for frenchname, englishname in translated_names.items():
            if(word == frenchname):
                return englishname

    workbook = openpyxl.load_workbook("./UserInput.xlsx")
    sheet    = workbook.active

    names = []
    for index in range(2, (sheet.max_row + 1)):
        names.append(french2english(sheet[f"A{index}"].value))

    patients_values = []
    for i in range(2, (sheet.max_column + 1)):
        column_letter = openpyxl.utils.get_column_letter(i)
        patient_values = []
        for j in range(2, (sheet.max_row + 1)):
            if(sheet[f"{column_letter}{j}"].value != None):
                patient_values.append(sheet[f"{column_letter}{j}"].value)
            
        patients_values.append(patient_values)

    with open("./UserInput.csv", "w", newline="") as file:
        writer = csv.writer(file)
        writer.writerow(names)
        for patient_values in patients_values:
            writer.writerow(patient_values)

    return best_knn_model.predict(pandas.read_csv("./UserInput.csv"))